"""
Excel Exporter Module
==================

This module provides functionality to export dataframes to nicely formatted Excel files.
It handles column styling, header formatting, and auto-filtering.
"""
import logging
from typing import Dict, List, Optional, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

def export_to_excel(
    df: pd.DataFrame,
    file_path: str,
    sheet_name: str = "DOGE Contracts",
    table_name: str = "DOGEContractsTable",
    number_format_columns: Optional[Dict[str, str]] = None,
    column_widths: Optional[Dict[str, int]] = None,
    freeze_panes: str = "A2"
) -> None:
    """
    Export a pandas DataFrame to a nicely formatted Excel file.
    
    Args:
        df: DataFrame to export
        file_path: Path to save the Excel file
        sheet_name: Name of the worksheet
        table_name: Name of the Excel table
        number_format_columns: Dict mapping column names to number formats
        column_widths: Dict mapping column names to column widths
        freeze_panes: Cell reference for freeze panes (e.g., 'A2')
    """
    logging.info(f"Exporting data to Excel file: {file_path}")
    
    # Create a Pandas Excel writer using openpyxl as the engine
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        # Convert the dataframe to an Excel object
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Get the workbook and the worksheet
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Apply formatting to the headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Apply header styling
        for col_num, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Apply number formatting to specified columns
        if number_format_columns:
            for col_name, number_format in number_format_columns.items():
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 1
                    col_letter = get_column_letter(col_idx)
                    
                    for row_num in range(2, len(df) + 2):  # +2 because of 1-indexing and header
                        cell = worksheet.cell(row=row_num, column=col_idx)
                        cell.number_format = number_format
        
        # Set column widths
        if not column_widths:
            # Default column widths based on data type
            column_widths = {}
            for i, column in enumerate(df.columns):
                if df[column].dtype in [int, float]:
                    column_widths[column] = 15
                elif "date" in column.lower():
                    column_widths[column] = 15
                else:
                    column_widths[column] = 30
        
        for col_name, width in column_widths.items():
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                worksheet.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Create a table
        table_ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
        table = Table(displayName=table_name, ref=table_ref)
        
        # Add a default style to the table
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        
        # Add the table to the worksheet
        worksheet.add_table(table)
        
        # Set freeze panes
        worksheet.freeze_panes = freeze_panes
        
    logging.info(f"Successfully exported data to {file_path} with formatting")

def export_doge_data_to_excel(df: pd.DataFrame, file_path: str) -> None:
    """
    Export DOGE contract data to a nicely formatted Excel file with 
    appropriate formatting for the specific columns.
    
    Args:
        df: DataFrame containing DOGE contract data
        file_path: Path to save the Excel file
    """
    # Define number format for currency columns
    number_formats = {
        "Total Contract Value (TCV)": '"$"#,##0.00_);("$"#,##0.00)',
        "Savings": '"$"#,##0.00_);("$"#,##0.00)'
    }
    
    # Define custom column widths
    column_widths = {
        "PIID": 20,
        "Buying Org 1": 25,
        "Buying Org 2": 25,
        "Buying Org 3": 25,
        "NAICS": 12,
        "PSC": 12,
        "Total Contract Value (TCV)": 20,
        "Savings": 20,
        "Deleted On": 15,
        "Incumbent": 30,
        "Description": 50,
        "Status": 20,
        "FPDS Link": 50
    }
    
    # Export to Excel with specified formatting
    export_to_excel(
        df=df,
        file_path=file_path,
        sheet_name="DOGE Contracts",
        table_name="DOGEContractsTable",
        number_format_columns=number_formats,
        column_widths=column_widths,
        freeze_panes="A2"
    )
